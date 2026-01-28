from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, Request
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import httpx
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 24 * 7  # 7 days

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ===== MODELS =====

class User(BaseModel):
    user_id: str
    email: str
    name: str
    role: str  # 'employee', 'manager', 'admin'
    created_at: datetime

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = 'employee'

class UserLogin(BaseModel):
    email: str
    password: str

class TokenResponse(BaseModel):
    token: str
    user: User

class Category(BaseModel):
    category_id: str
    name: str
    description: Optional[str] = None
    created_at: datetime

class CategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None

class ProductType(BaseModel):
    type_id: str
    name: str
    category_id: str
    category_name: Optional[str] = None
    created_at: datetime

class ProductTypeCreate(BaseModel):
    name: str
    category_id: str

class Product(BaseModel):
    product_id: str
    sku: str
    name: str
    category_id: str
    category_name: Optional[str] = None
    type_id: str
    type_name: Optional[str] = None
    status: str  # 'active', 'inactive'
    current_stock: Optional[int] = 0
    created_at: datetime

class ProductCreate(BaseModel):
    sku: str
    name: str
    category_id: str
    type_id: str
    status: str = 'active'

class Supplier(BaseModel):
    supplier_id: str
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    has_vat: bool = False
    created_at: datetime

class SupplierCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    has_vat: bool = False

class PurchaseOrderItem(BaseModel):
    product_id: str
    product_name: Optional[str] = None
    quantity: int
    unit_price: float
    total: float

class PurchaseOrder(BaseModel):
    po_id: str
    supplier_id: str
    supplier_name: Optional[str] = None
    date: datetime
    items: List[PurchaseOrderItem]
    total_amount: float
    payment_status: str  # 'unpaid', 'partial', 'paid'
    paid_amount: float = 0
    created_by: str
    created_at: datetime

class PurchaseOrderCreate(BaseModel):
    supplier_id: str
    date: datetime
    items: List[PurchaseOrderItem]

class Customer(BaseModel):
    customer_id: str
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    group: Optional[str] = None
    notes: Optional[str] = None
    created_at: datetime

class CustomerCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    group: Optional[str] = None
    notes: Optional[str] = None

class SalesOrderItem(BaseModel):
    product_id: str
    product_name: Optional[str] = None
    quantity: int
    unit_price: float
    total: float

class SalesOrder(BaseModel):
    order_id: str
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    date: datetime
    order_type: str  # 'normal', 'livestream'
    items: List[SalesOrderItem]
    total_amount: float
    payment_status: str  # 'unpaid', 'partial', 'paid'
    paid_amount: float = 0
    created_by: str
    created_at: datetime

class SalesOrderCreate(BaseModel):
    customer_id: Optional[str] = None
    date: datetime
    order_type: str = 'normal'
    items: List[SalesOrderItem]

class InventoryItem(BaseModel):
    product_id: str
    product_name: Optional[str] = None
    sku: Optional[str] = None
    quantity: int
    last_updated: datetime

class Transaction(BaseModel):
    transaction_id: str
    date: datetime
    type: str  # 'income', 'expense'
    category: str  # 'sales', 'purchase', 'other'
    amount: float
    description: Optional[str] = None
    related_to: Optional[str] = None
    created_by: str
    created_at: datetime

class TransactionCreate(BaseModel):
    date: datetime
    type: str
    category: str
    amount: float
    description: Optional[str] = None
    related_to: Optional[str] = None

class DashboardStats(BaseModel):
    total_revenue: float
    total_expenses: float
    total_profit: float
    total_orders: int
    pending_orders: int
    low_stock_products: int
    total_customers: int
    total_suppliers: int

class ReportQuery(BaseModel):
    start_date: datetime
    end_date: datetime
    report_type: str  # 'sales', 'purchases', 'inventory', 'cashflow'

# ===== AUTH HELPER =====

async def get_current_user(request: Request) -> User:
    # Check cookie first
    token = request.cookies.get('session_token')
    
    # Fallback to Authorization header
    if not token:
        auth_header = request.headers.get('Authorization')
        if auth_header and auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
    
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    
    try:
        # Try JWT token first
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user_id = payload.get('user_id')
            if not user_id:
                raise HTTPException(status_code=401, detail="Invalid token")
        except jwt.InvalidTokenError:
            # Try Google OAuth session
            session_doc = await db.user_sessions.find_one({'session_token': token}, {'_id': 0})
            if not session_doc:
                raise HTTPException(status_code=401, detail="Session not found")
            
            expires_at = session_doc['expires_at']
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            
            if expires_at < datetime.now(timezone.utc):
                raise HTTPException(status_code=401, detail="Session expired")
            
            user_id = session_doc['user_id']
        
        user_doc = await db.users.find_one({'user_id': user_id}, {'_id': 0})
        if not user_doc:
            raise HTTPException(status_code=401, detail="User not found")
        
        if isinstance(user_doc['created_at'], str):
            user_doc['created_at'] = datetime.fromisoformat(user_doc['created_at'])
        
        return User(**user_doc)
    except Exception as e:
        logger.error(f"Auth error: {e}")
        raise HTTPException(status_code=401, detail="Invalid authentication")

async def require_role(user: User, allowed_roles: List[str]):
    if user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

# ===== AUTH ROUTES =====

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    # Check if user exists
    existing = await db.users.find_one({'email': user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Hash password
    password_hash = bcrypt.hashpw(user_data.password.encode('utf-8'), bcrypt.gensalt())
    
    # Create user
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    user_doc = {
        'user_id': user_id,
        'email': user_data.email,
        'name': user_data.name,
        'password_hash': password_hash.decode('utf-8'),
        'role': user_data.role,
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Create token
    token = jwt.encode(
        {'user_id': user_id, 'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)},
        JWT_SECRET,
        algorithm=JWT_ALGORITHM
    )
    
    user = User(
        user_id=user_id,
        email=user_data.email,
        name=user_data.name,
        role=user_data.role,
        created_at=datetime.fromisoformat(user_doc['created_at'])
    )
    
    return TokenResponse(token=token, user=user)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({'email': credentials.email})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Verify password
    if not bcrypt.checkpw(credentials.password.encode('utf-8'), user_doc['password_hash'].encode('utf-8')):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Create token
    token = jwt.encode(
        {'user_id': user_doc['user_id'], 'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)},
        JWT_SECRET,
        algorithm=JWT_ALGORITHM
    )
    
    user = User(
        user_id=user_doc['user_id'],
        email=user_doc['email'],
        name=user_doc['name'],
        role=user_doc['role'],
        created_at=datetime.fromisoformat(user_doc['created_at']) if isinstance(user_doc['created_at'], str) else user_doc['created_at']
    )
    
    return TokenResponse(token=token, user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(user: User = Depends(get_current_user)):
    return user

@api_router.post("/auth/google/session")
async def google_auth_session(request: Request, response: Response):
    session_id = request.headers.get('X-Session-ID')
    if not session_id:
        raise HTTPException(status_code=400, detail="Missing session ID")
    
    # Call Emergent Auth API
    async with httpx.AsyncClient() as client:
        resp = await client.get(
            'https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data',
            headers={'X-Session-ID': session_id}
        )
        if resp.status_code != 200:
            raise HTTPException(status_code=401, detail="Invalid session")
        
        data = resp.json()
    
    # Create or update user
    user_doc = await db.users.find_one({'email': data['email']}, {'_id': 0})
    if user_doc:
        user_id = user_doc['user_id']
        await db.users.update_one(
            {'user_id': user_id},
            {'$set': {'name': data['name'], 'picture': data.get('picture')}}
        )
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_doc = {
            'user_id': user_id,
            'email': data['email'],
            'name': data['name'],
            'picture': data.get('picture'),
            'role': 'employee',
            'created_at': datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(user_doc)
    
    # Store session
    session_token = data['session_token']
    await db.user_sessions.insert_one({
        'user_id': user_id,
        'session_token': session_token,
        'expires_at': datetime.now(timezone.utc) + timedelta(days=7),
        'created_at': datetime.now(timezone.utc)
    })
    
    # Set cookie
    response.set_cookie(
        key='session_token',
        value=session_token,
        httponly=True,
        secure=True,
        samesite='none',
        path='/',
        max_age=7 * 24 * 60 * 60
    )
    
    user = User(
        user_id=user_id,
        email=data['email'],
        name=data['name'],
        role=user_doc.get('role', 'employee'),
        created_at=datetime.fromisoformat(user_doc['created_at']) if isinstance(user_doc.get('created_at'), str) else user_doc.get('created_at', datetime.now(timezone.utc))
    )
    
    return {'user': user}

@api_router.post("/auth/logout")
async def logout(response: Response, user: User = Depends(get_current_user)):
    # Delete session from DB
    await db.user_sessions.delete_many({'user_id': user.user_id})
    
    # Clear cookie
    response.delete_cookie('session_token', path='/')
    
    return {'message': 'Logged out successfully'}

# ===== CATEGORY ROUTES =====

@api_router.get("/categories", response_model=List[Category])
async def get_categories(user: User = Depends(get_current_user)):
    categories = await db.categories.find({}, {'_id': 0}).to_list(1000)
    for cat in categories:
        if isinstance(cat['created_at'], str):
            cat['created_at'] = datetime.fromisoformat(cat['created_at'])
    return categories

@api_router.post("/categories", response_model=Category)
async def create_category(category_data: CategoryCreate, user: User = Depends(get_current_user)):
    await require_role(user, ['manager', 'admin'])
    
    category_id = f"cat_{uuid.uuid4().hex[:12]}"
    category_doc = {
        'category_id': category_id,
        'name': category_data.name,
        'description': category_data.description,
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    await db.categories.insert_one(category_doc)
    
    return Category(
        category_id=category_id,
        name=category_data.name,
        description=category_data.description,
        created_at=datetime.fromisoformat(category_doc['created_at'])
    )

@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(category_id: str, category_data: CategoryCreate, user: User = Depends(get_current_user)):
    await require_role(user, ['manager', 'admin'])
    
    result = await db.categories.update_one(
        {'category_id': category_id},
        {'$set': {'name': category_data.name, 'description': category_data.description}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    category_doc = await db.categories.find_one({'category_id': category_id}, {'_id': 0})
    if isinstance(category_doc['created_at'], str):
        category_doc['created_at'] = datetime.fromisoformat(category_doc['created_at'])
    return Category(**category_doc)

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, user: User = Depends(get_current_user)):
    await require_role(user, ['manager', 'admin'])
    
    result = await db.categories.delete_one({'category_id': category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    
    return {'message': 'Category deleted successfully'}

# ===== PRODUCT TYPE ROUTES =====

@api_router.get("/product-types", response_model=List[ProductType])
async def get_product_types(user: User = Depends(get_current_user)):
    types = await db.product_types.find({}, {'_id': 0}).to_list(1000)
    for pt in types:
        if isinstance(pt['created_at'], str):
            pt['created_at'] = datetime.fromisoformat(pt['created_at'])
        # Get category name
        cat = await db.categories.find_one({'category_id': pt['category_id']}, {'_id': 0})
        pt['category_name'] = cat['name'] if cat else None
    return types

@api_router.post("/product-types", response_model=ProductType)
async def create_product_type(type_data: ProductTypeCreate, user: User = Depends(get_current_user)):
    await require_role(user, ['manager', 'admin'])
    
    type_id = f"type_{uuid.uuid4().hex[:12]}"
    type_doc = {
        'type_id': type_id,
        'name': type_data.name,
        'category_id': type_data.category_id,
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    await db.product_types.insert_one(type_doc)
    
    cat = await db.categories.find_one({'category_id': type_data.category_id}, {'_id': 0})
    
    return ProductType(
        type_id=type_id,
        name=type_data.name,
        category_id=type_data.category_id,
        category_name=cat['name'] if cat else None,
        created_at=datetime.fromisoformat(type_doc['created_at'])
    )

# ===== PRODUCT ROUTES =====

@api_router.get("/products", response_model=List[Product])
async def get_products(user: User = Depends(get_current_user)):
    products = await db.products.find({}, {'_id': 0}).to_list(1000)
    for prod in products:
        if isinstance(prod['created_at'], str):
            prod['created_at'] = datetime.fromisoformat(prod['created_at'])
        
        # Get category and type names
        cat = await db.categories.find_one({'category_id': prod['category_id']}, {'_id': 0})
        prod['category_name'] = cat['name'] if cat else None
        
        ptype = await db.product_types.find_one({'type_id': prod['type_id']}, {'_id': 0})
        prod['type_name'] = ptype['name'] if ptype else None
        
        # Get current stock
        inventory = await db.inventory.find_one({'product_id': prod['product_id']}, {'_id': 0})
        prod['current_stock'] = inventory['quantity'] if inventory else 0
    
    return products

@api_router.post("/products", response_model=Product)
async def create_product(product_data: ProductCreate, user: User = Depends(get_current_user)):
    await require_role(user, ['manager', 'admin'])
    
    # Check if SKU already exists
    existing = await db.products.find_one({'sku': product_data.sku})
    if existing:
        raise HTTPException(status_code=400, detail="SKU already exists")
    
    product_id = f"prod_{uuid.uuid4().hex[:12]}"
    product_doc = {
        'product_id': product_id,
        'sku': product_data.sku,
        'name': product_data.name,
        'category_id': product_data.category_id,
        'type_id': product_data.type_id,
        'status': product_data.status,
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    await db.products.insert_one(product_doc)
    
    # Initialize inventory
    await db.inventory.insert_one({
        'product_id': product_id,
        'quantity': 0,
        'last_updated': datetime.now(timezone.utc)
    })
    
    cat = await db.categories.find_one({'category_id': product_data.category_id}, {'_id': 0})
    ptype = await db.product_types.find_one({'type_id': product_data.type_id}, {'_id': 0})
    
    return Product(
        product_id=product_id,
        sku=product_data.sku,
        name=product_data.name,
        category_id=product_data.category_id,
        category_name=cat['name'] if cat else None,
        type_id=product_data.type_id,
        type_name=ptype['name'] if ptype else None,
        status=product_data.status,
        current_stock=0,
        created_at=datetime.fromisoformat(product_doc['created_at'])
    )

@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(product_id: str, product_data: ProductCreate, user: User = Depends(get_current_user)):
    await require_role(user, ['manager', 'admin'])
    
    result = await db.products.update_one(
        {'product_id': product_id},
        {'$set': {
            'sku': product_data.sku,
            'name': product_data.name,
            'category_id': product_data.category_id,
            'type_id': product_data.type_id,
            'status': product_data.status
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    
    product_doc = await db.products.find_one({'product_id': product_id}, {'_id': 0})
    if isinstance(product_doc['created_at'], str):
        product_doc['created_at'] = datetime.fromisoformat(product_doc['created_at'])
    
    cat = await db.categories.find_one({'category_id': product_doc['category_id']}, {'_id': 0})
    ptype = await db.product_types.find_one({'type_id': product_doc['type_id']}, {'_id': 0})
    inventory = await db.inventory.find_one({'product_id': product_id}, {'_id': 0})
    
    product_doc['category_name'] = cat['name'] if cat else None
    product_doc['type_name'] = ptype['name'] if ptype else None
    product_doc['current_stock'] = inventory['quantity'] if inventory else 0
    
    return Product(**product_doc)

# ===== SUPPLIER ROUTES =====

@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers(user: User = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {'_id': 0}).to_list(1000)
    for sup in suppliers:
        if isinstance(sup['created_at'], str):
            sup['created_at'] = datetime.fromisoformat(sup['created_at'])
    return suppliers

@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier_data: SupplierCreate, user: User = Depends(get_current_user)):
    await require_role(user, ['manager', 'admin'])
    
    supplier_id = f"sup_{uuid.uuid4().hex[:12]}"
    supplier_doc = {
        'supplier_id': supplier_id,
        'name': supplier_data.name,
        'contact_person': supplier_data.contact_person,
        'phone': supplier_data.phone,
        'email': supplier_data.email,
        'address': supplier_data.address,
        'has_vat': supplier_data.has_vat,
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    await db.suppliers.insert_one(supplier_doc)
    
    return Supplier(**{**supplier_doc, 'created_at': datetime.fromisoformat(supplier_doc['created_at'])})

# ===== PURCHASE ORDER ROUTES =====

@api_router.get("/purchase-orders", response_model=List[PurchaseOrder])
async def get_purchase_orders(user: User = Depends(get_current_user)):
    orders = await db.purchase_orders.find({}, {'_id': 0}).to_list(1000)
    for order in orders:
        if isinstance(order['date'], str):
            order['date'] = datetime.fromisoformat(order['date'])
        if isinstance(order['created_at'], str):
            order['created_at'] = datetime.fromisoformat(order['created_at'])
        
        # Get supplier name
        supplier = await db.suppliers.find_one({'supplier_id': order['supplier_id']}, {'_id': 0})
        order['supplier_name'] = supplier['name'] if supplier else None
    
    return orders

@api_router.post("/purchase-orders", response_model=PurchaseOrder)
async def create_purchase_order(order_data: PurchaseOrderCreate, user: User = Depends(get_current_user)):
    await require_role(user, ['manager', 'admin'])
    
    # Calculate totals
    total_amount = sum(item.quantity * item.unit_price for item in order_data.items)
    
    # Get product names
    items_with_names = []
    for item in order_data.items:
        product = await db.products.find_one({'product_id': item.product_id}, {'_id': 0})
        items_with_names.append({
            'product_id': item.product_id,
            'product_name': product['name'] if product else None,
            'quantity': item.quantity,
            'unit_price': item.unit_price,
            'total': item.quantity * item.unit_price
        })
    
    po_id = f"po_{uuid.uuid4().hex[:12]}"
    po_doc = {
        'po_id': po_id,
        'supplier_id': order_data.supplier_id,
        'date': order_data.date.isoformat(),
        'items': items_with_names,
        'total_amount': total_amount,
        'payment_status': 'unpaid',
        'paid_amount': 0,
        'created_by': user.user_id,
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    await db.purchase_orders.insert_one(po_doc)
    
    # Update inventory
    for item in order_data.items:
        await db.inventory.update_one(
            {'product_id': item.product_id},
            {'$inc': {'quantity': item.quantity}, '$set': {'last_updated': datetime.now(timezone.utc)}},
            upsert=True
        )
    
    # Create transaction
    await db.transactions.insert_one({
        'transaction_id': f"txn_{uuid.uuid4().hex[:12]}",
        'date': order_data.date,
        'type': 'expense',
        'category': 'purchase',
        'amount': total_amount,
        'description': f"Purchase order {po_id}",
        'related_to': po_id,
        'created_by': user.user_id,
        'created_at': datetime.now(timezone.utc)
    })
    
    supplier = await db.suppliers.find_one({'supplier_id': order_data.supplier_id}, {'_id': 0})
    
    return PurchaseOrder(
        po_id=po_id,
        supplier_id=order_data.supplier_id,
        supplier_name=supplier['name'] if supplier else None,
        date=order_data.date,
        items=[PurchaseOrderItem(**item) for item in items_with_names],
        total_amount=total_amount,
        payment_status='unpaid',
        paid_amount=0,
        created_by=user.user_id,
        created_at=datetime.fromisoformat(po_doc['created_at'])
    )

# ===== CUSTOMER ROUTES =====

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(user: User = Depends(get_current_user)):
    customers = await db.customers.find({}, {'_id': 0}).to_list(1000)
    for cust in customers:
        if isinstance(cust['created_at'], str):
            cust['created_at'] = datetime.fromisoformat(cust['created_at'])
    return customers

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, user: User = Depends(get_current_user)):
    customer_id = f"cust_{uuid.uuid4().hex[:12]}"
    customer_doc = {
        'customer_id': customer_id,
        'name': customer_data.name,
        'phone': customer_data.phone,
        'email': customer_data.email,
        'address': customer_data.address,
        'group': customer_data.group,
        'notes': customer_data.notes,
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    await db.customers.insert_one(customer_doc)
    
    return Customer(**{**customer_doc, 'created_at': datetime.fromisoformat(customer_doc['created_at'])})

# ===== SALES ORDER ROUTES =====

@api_router.get("/sales-orders", response_model=List[SalesOrder])
async def get_sales_orders(user: User = Depends(get_current_user)):
    orders = await db.sales_orders.find({}, {'_id': 0}).to_list(1000)
    for order in orders:
        if isinstance(order['date'], str):
            order['date'] = datetime.fromisoformat(order['date'])
        if isinstance(order['created_at'], str):
            order['created_at'] = datetime.fromisoformat(order['created_at'])
        
        if order.get('customer_id'):
            customer = await db.customers.find_one({'customer_id': order['customer_id']}, {'_id': 0})
            order['customer_name'] = customer['name'] if customer else None
    
    return orders

@api_router.post("/sales-orders", response_model=SalesOrder)
async def create_sales_order(order_data: SalesOrderCreate, user: User = Depends(get_current_user)):
    # Check stock availability
    for item in order_data.items:
        inventory = await db.inventory.find_one({'product_id': item.product_id}, {'_id': 0})
        current_stock = inventory['quantity'] if inventory else 0
        if current_stock < item.quantity:
            product = await db.products.find_one({'product_id': item.product_id}, {'_id': 0})
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock for {product['name'] if product else item.product_id}. Available: {current_stock}"
            )
    
    # Calculate totals
    total_amount = sum(item.quantity * item.unit_price for item in order_data.items)
    
    # Get product names
    items_with_names = []
    for item in order_data.items:
        product = await db.products.find_one({'product_id': item.product_id}, {'_id': 0})
        items_with_names.append({
            'product_id': item.product_id,
            'product_name': product['name'] if product else None,
            'quantity': item.quantity,
            'unit_price': item.unit_price,
            'total': item.quantity * item.unit_price
        })
    
    order_id = f"so_{uuid.uuid4().hex[:12]}"
    so_doc = {
        'order_id': order_id,
        'customer_id': order_data.customer_id,
        'date': order_data.date.isoformat(),
        'order_type': order_data.order_type,
        'items': items_with_names,
        'total_amount': total_amount,
        'payment_status': 'unpaid',
        'paid_amount': 0,
        'created_by': user.user_id,
        'created_at': datetime.now(timezone.utc).isoformat()
    }
    await db.sales_orders.insert_one(so_doc)
    
    # Update inventory
    for item in order_data.items:
        await db.inventory.update_one(
            {'product_id': item.product_id},
            {'$inc': {'quantity': -item.quantity}, '$set': {'last_updated': datetime.now(timezone.utc)}}
        )
    
    # Create transaction
    await db.transactions.insert_one({
        'transaction_id': f"txn_{uuid.uuid4().hex[:12]}",
        'date': order_data.date,
        'type': 'income',
        'category': 'sales',
        'amount': total_amount,
        'description': f"Sales order {order_id}",
        'related_to': order_id,
        'created_by': user.user_id,
        'created_at': datetime.now(timezone.utc)
    })
    
    customer_name = None
    if order_data.customer_id:
        customer = await db.customers.find_one({'customer_id': order_data.customer_id}, {'_id': 0})
        customer_name = customer['name'] if customer else None
    
    return SalesOrder(
        order_id=order_id,
        customer_id=order_data.customer_id,
        customer_name=customer_name,
        date=order_data.date,
        order_type=order_data.order_type,
        items=[SalesOrderItem(**item) for item in items_with_names],
        total_amount=total_amount,
        payment_status='unpaid',
        paid_amount=0,
        created_by=user.user_id,
        created_at=datetime.fromisoformat(so_doc['created_at'])
    )

# ===== INVENTORY ROUTES =====

@api_router.get("/inventory", response_model=List[InventoryItem])
async def get_inventory(user: User = Depends(get_current_user)):
    inventory = await db.inventory.find({}, {'_id': 0}).to_list(1000)
    for item in inventory:
        product = await db.products.find_one({'product_id': item['product_id']}, {'_id': 0})
        item['product_name'] = product['name'] if product else None
        item['sku'] = product['sku'] if product else None
        if isinstance(item['last_updated'], str):
            item['last_updated'] = datetime.fromisoformat(item['last_updated'])
    return inventory

# ===== DASHBOARD ROUTES =====

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(user: User = Depends(get_current_user)):
    # Calculate revenue
    income_txns = await db.transactions.find({'type': 'income'}, {'_id': 0}).to_list(10000)
    total_revenue = sum(txn['amount'] for txn in income_txns)
    
    # Calculate expenses
    expense_txns = await db.transactions.find({'type': 'expense'}, {'_id': 0}).to_list(10000)
    total_expenses = sum(txn['amount'] for txn in expense_txns)
    
    # Calculate profit
    total_profit = total_revenue - total_expenses
    
    # Count orders
    total_orders = await db.sales_orders.count_documents({})
    pending_orders = await db.sales_orders.count_documents({'payment_status': {'$ne': 'paid'}})
    
    # Low stock products (< 10)
    low_stock = await db.inventory.count_documents({'quantity': {'$lt': 10}})
    
    # Count customers and suppliers
    total_customers = await db.customers.count_documents({})
    total_suppliers = await db.suppliers.count_documents({})
    
    return DashboardStats(
        total_revenue=total_revenue,
        total_expenses=total_expenses,
        total_profit=total_profit,
        total_orders=total_orders,
        pending_orders=pending_orders,
        low_stock_products=low_stock,
        total_customers=total_customers,
        total_suppliers=total_suppliers
    )

# ===== REPORTS & EXPORT =====

@api_router.post("/reports/export")
async def export_report(query: ReportQuery, user: User = Depends(get_current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = query.report_type.capitalize()
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="24A853", end_color="24A853", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    if query.report_type == 'sales':
        ws.append(['Order ID', 'Date', 'Customer', 'Type', 'Total Amount', 'Payment Status'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        orders = await db.sales_orders.find({
            'date': {
                '$gte': query.start_date.isoformat(),
                '$lte': query.end_date.isoformat()
            }
        }, {'_id': 0}).to_list(10000)
        
        for order in orders:
            customer_name = ''
            if order.get('customer_id'):
                customer = await db.customers.find_one({'customer_id': order['customer_id']}, {'_id': 0})
                customer_name = customer['name'] if customer else ''
            
            ws.append([
                order['order_id'],
                order['date'],
                customer_name,
                order['order_type'],
                order['total_amount'],
                order['payment_status']
            ])
    
    elif query.report_type == 'purchases':
        ws.append(['PO ID', 'Date', 'Supplier', 'Total Amount', 'Payment Status'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        orders = await db.purchase_orders.find({
            'date': {
                '$gte': query.start_date.isoformat(),
                '$lte': query.end_date.isoformat()
            }
        }, {'_id': 0}).to_list(10000)
        
        for order in orders:
            supplier = await db.suppliers.find_one({'supplier_id': order['supplier_id']}, {'_id': 0})
            
            ws.append([
                order['po_id'],
                order['date'],
                supplier['name'] if supplier else '',
                order['total_amount'],
                order['payment_status']
            ])
    
    elif query.report_type == 'inventory':
        ws.append(['Product ID', 'SKU', 'Product Name', 'Quantity', 'Last Updated'])
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        inventory = await db.inventory.find({}, {'_id': 0}).to_list(10000)
        
        for item in inventory:
            product = await db.products.find_one({'product_id': item['product_id']}, {'_id': 0})
            
            ws.append([
                item['product_id'],
                product['sku'] if product else '',
                product['name'] if product else '',
                item['quantity'],
                item['last_updated']
            ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"{query.report_type}_report_{query.start_date.date()}_{query.end_date.date()}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ===== USER MANAGEMENT (Admin only) =====

@api_router.get("/users", response_model=List[User])
async def get_users(user: User = Depends(get_current_user)):
    await require_role(user, ['admin'])
    
    users = await db.users.find({}, {'_id': 0, 'password_hash': 0}).to_list(1000)
    for u in users:
        if isinstance(u['created_at'], str):
            u['created_at'] = datetime.fromisoformat(u['created_at'])
    return users

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role: str, user: User = Depends(get_current_user)):
    await require_role(user, ['admin'])
    
    if role not in ['employee', 'manager', 'admin']:
        raise HTTPException(status_code=400, detail="Invalid role")
    
    result = await db.users.update_one(
        {'user_id': user_id},
        {'$set': {'role': role}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {'message': 'User role updated successfully'}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
